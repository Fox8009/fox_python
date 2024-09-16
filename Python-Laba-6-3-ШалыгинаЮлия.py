# Использовать библиотеку openpyxl
# Создать документ Excel с одним рабочим листом. Назвать его ‘Python’.
# Столбец A заполнить числами от 3 до 15.
# Столбец B заполнить следующим образом: каждое третье значение столбца А умножается на 2, остальные просто переносятся из А.
# По данным столбца B построить BarChart и  PieChart.
# Сохранить

import openpyxl


# создание документа и листа
from openpyxl.chart import BarChart, PieChart, Reference


def createChart(chart, title, height, data):
    chart.title = title
    chart.height = height
    chart.add_data(data, titles_from_data=False)
    return chart


wb = openpyxl.Workbook()
wb.create_sheet(title='Python', index=0)
sheet = wb['Python']
# заполнение таблицы по условиям
for col in range(1, 3):
    for row in range(1, 14):
        cell = sheet.cell(row=row, column=col)
        value = row + 2
        cell.value = value*2 if row % 3 == 0 and col == 2 else value

data = Reference(sheet, min_col=2, min_row=1, max_col=2, max_row=13)

# создание диаграмм
sheet.add_chart(createChart(BarChart(), 'Bars', 11, data), 'D2')
sheet.add_chart(createChart(PieChart(), 'Pies', 11, data), 'N2')

wb.save('laba6.xlsx')




